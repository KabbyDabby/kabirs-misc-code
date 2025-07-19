import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import anthropic
import PyPDF2
import re
from pathlib import Path
import json
import time


class CDSAnalyzer:
    def __init__(self, api_key, pdf_folder_path):
        self.client = anthropic.Anthropic(api_key=api_key)
        self.pdf_folder_path = pdf_folder_path
        self.results = []
        self.tokens_used = 0
        self.last_request_time = time.time()

        # Define color fills
        self.colors = {
            "green": PatternFill(
                start_color="00FF00", end_color="00FF00", fill_type="solid"
            ),
            "yellow": PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            ),
            "red": PatternFill(
                start_color="FF0000", end_color="FF0000", fill_type="solid"
            ),
        }

    def extract_text_from_pdf(self, pdf_path):
        """Extract text from PDF file"""
        try:
            with open(pdf_path, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    try:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
                    except Exception as e:
                        print(
                            f"Warning: Error extracting text from page in {pdf_path}: {e}"
                        )
                        continue

                if not text.strip():
                    print(f"Warning: No text extracted from {pdf_path}")
                    return None

                return text
        except Exception as e:
            print(f"Error reading PDF {pdf_path}: {e}")
            return None

    def analyze_cds_with_claude(self, pdf_text, university_name):
        """Use Claude API to analyze CDS data"""
        prompt = f"""
        Analyze this Common Data Set (CDS) for {university_name} and extract the following information. 
        Return the data in JSON format with these exact keys:

        {{
            "university_name": "{university_name}",
            "institution_type": "public/private_nonprofit/proprietary",
            "data_year": "YYYY (e.g., 2025 for 2024-25 data)",
            "percent_female_undergrad": "percentage as decimal (e.g., 0.52 for 52%)",
            "total_undergrads": "number",
            "total_grads": "number",
            "undergrad_grad_ratio": "decimal ratio",
            "percent_asian_undergrads": "percentage as decimal",
            "four_year_graduation_rate": "percentage as decimal",
            "retention_rate": "percentage as decimal",
            "male_acceptance_rate": "percentage as decimal",
            "female_acceptance_rate": "percentage as decimal",
            "net_acceptance_rate": "percentage as decimal",
            "instate_acceptance_rate": "percentage as decimal",
            "outstate_acceptance_rate": "percentage as decimal",
            "male_yield_rate": "percentage as decimal",
            "female_yield_rate": "percentage as decimal",
            "net_yield_rate": "percentage as decimal",
            "instate_yield_rate": "percentage as decimal",
            "outstate_yield_rate": "percentage as decimal",
            "sat_act_policy": "required/optional/blind",
            "sat_composite_50th": "number or null",
            "sat_math_50th": "number or null",
            "sat_reading_50th": "number or null",
            "sat_composite_75th": "number or null",
            "sat_math_75th": "number or null",
            "sat_reading_75th": "number or null",
            "early_action_available": "yes/no",
            "early_decision_available": "yes/no",
            "ed_acceptance_rate": "percentage as decimal or null",
            "ea_acceptance_rate": "percentage as decimal or null",
            "double_majoring_allowed": "yes/no",
            "tuition_room_board": "total cost number"
        }}

        Key sections to look for:
        - Section A: General Information (institution type, year)
        - Section B: Enrollment and Persistence (graduation rates, retention, demographics)
        - Section C: First-Time Applicants (admission data, SAT/ACT scores)
        - Section G: Annual Expenses (tuition and fees)

               
        
        Important instructions:
        1. For institution type: Look in Section A for institution type
        2. For data year: Look for the academic year in Section A (e.g., 2024-25 should be recorded as 2025)
        3. For female percentage: Use first-time, degree-seeking undergraduate data from Section B
        4. For total undergrads and grads: Use Section B totals
        5. For acceptance rates: Use Section C data (admitted/applied)
        6. For yield rates: Use Section C data (enrolled/admitted)
        7. For SAT/ACT policy: Look in Section C for test policy
        8. For SAT scores: Use Section C-1 for test scores
        9. For early action/decision: Look in Section C for early admission policies
        10. For double majoring: Look in Section H for academic policies
        11. For tuition: Use Section G for total cost (tuition + room and board)

        If any data is not available, use null for numbers and "unknown" for text fields.
        Make sure to calculate all percentages as decimals (e.g., 0.52 for 52%).


        For calculations:
        - Percent female undergrad: Use first-time, degree-seeking undergraduate data
        - Acceptance rates: admits/applicants
        - Yield rates: enrolled/admits
        - Graduation rate: Use 4-year rate from cohort data (B8/B5)
        - Asian percentage: From first-year demographic data

        CDS Text:
        {pdf_text[:50000]}  # Limit text to avoid token limits
        """

        try:
            response = self.client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4000,
                messages=[{"role": "user", "content": prompt}],
            )

            # Calculate tokens used if available
            if hasattr(response, "usage") and hasattr(response.usage, "total_tokens"):
                self.tokens_used += response.usage.total_tokens

            # Check if tokens used per minute exceeds 20k
            current_time = time.time()
            time_diff = current_time - self.last_request_time
            if time_diff < 60 and self.tokens_used > 20000:
                wait_time = 60 - time_diff
                print(
                    f"Waiting for {wait_time:.2f} seconds to avoid exceeding token limit..."
                )
                time.sleep(wait_time)
                self.tokens_used = 0
                self.last_request_time = time.time()

            # Extract JSON from response
            response_text = response.content[0].text
            json_start = response_text.find("{")
            json_end = response_text.rfind("}") + 1
            json_str = response_text[json_start:json_end]

            return json.loads(json_str)

        except Exception as e:
            print(f"Error analyzing CDS for {university_name}: {e}")
            return None

    def get_color_for_value(self, value, criteria):
        """Determine color based on criteria"""
        if criteria == "institution_type":
            if value == "public":
                return "yellow"
            elif value == "private_nonprofit":
                return "green"
            elif value == "proprietary":
                return "red"

        elif criteria == "data_year":
            if value == 2025:
                return "green"
            elif value == 2024:
                return "yellow"
            else:
                return "red"

        elif criteria == "percent_female":
            abs_diff = abs(value - 0.5)
            if abs_diff < 0.05:
                return "green"
            elif abs_diff <= 0.10:
                return "yellow"
            else:
                return "red"

        elif criteria == "undergrad_grad_ratio":
            if value > 1:
                return "green"
            elif value > 0.75:
                return "yellow"
            else:
                return "red"

        elif criteria == "acceptance_rate":
            if value < 0.10:
                return "red"
            elif value < 0.20:
                return "yellow"
            else:
                return "green"

        elif criteria == "yield_rate":
            if value > 0.60:
                return "green"
            elif value > 0.40:
                return "yellow"
            else:
                return "red"

        elif criteria == "sat_act_policy":
            if value == "required":
                return "green"
            elif value == "optional":
                return "yellow"
            elif value == "blind":
                return "red"

        elif criteria == "sat_score":
            if value > 1520:
                return "red"
            elif value > 1400:
                return "yellow"
            else:
                return "red"

        elif criteria == "sat_section":
            if value > 760:
                return "red"
            elif value > 700:
                return "yellow"
            else:
                return "red"

        elif criteria == "yes_no_green":
            return "green" if value == "yes" else "red"

        elif criteria == "tuition":
            if value > 50000:
                return "red"
            elif value > 10000:
                return "yellow"
            else:
                return "green"

        return None

    def process_pdfs(self):
        """Process all PDF files in the folder"""
        pdf_files = list(Path(self.pdf_folder_path).glob("*.pdf"))

        for pdf_file in pdf_files:
            print(f"Processing {pdf_file.name}...")

            # Extract university name from filename
            university_name = pdf_file.stem.replace(" CDS", "")

            # Extract text from PDF
            pdf_text = self.extract_text_from_pdf(pdf_file)
            if not pdf_text:
                continue

            # Add longer delay between requests to respect rate limits
            time.sleep(5)  # Wait 5 seconds between each PDF processing

            # Analyze with Claude
            max_retries = 3
            retry_delay = 5  # Initial delay in seconds

            for attempt in range(max_retries):
                try:
                    analysis = self.analyze_cds_with_claude(pdf_text, university_name)
                    if analysis:
                        self.results.append(analysis)
                        print(f"Successfully analyzed {university_name}")
                        break
                except Exception as e:
                    if "rate_limit_error" in str(e) and attempt < max_retries - 1:
                        print(
                            f"Rate limit hit, waiting {retry_delay} seconds before retry..."
                        )
                        time.sleep(retry_delay)
                        retry_delay *= 2  # Exponential backoff
                    else:
                        print(f"Error analyzing CDS for {university_name}: {e}")
                        break

    def create_excel_file(self, output_path="university_cds_analysis.xlsx"):
        """Create Excel file with formatting and color coding"""
        if not self.results:
            print("No results to export")
            return

        # Delete existing file if it exists
        if os.path.exists(output_path):
            os.remove(output_path)

        # Create DataFrame
        df = pd.DataFrame(self.results)

        # Create Excel writer
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="CDS Analysis", index=False)

            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets["CDS Analysis"]

            # Helper function to safely convert to float
            def safe_float(value, default=0):
                if value in [None, "null", "unknown", ""]:
                    return default
                try:
                    return float(value)
                except (ValueError, TypeError):
                    return default

            # Helper function to safely convert to int
            def safe_int(value, default=0):
                if value in [None, "null", "unknown", ""]:
                    return default
                try:
                    return int(value)
                except (ValueError, TypeError):
                    return default

            # Apply formatting and colors
            for row_idx, row_data in enumerate(
                self.results, start=2
            ):  # Start from 2 (after header)
                col_idx = 1

                # University name (no color)
                col_idx += 1

                # Institution type
                inst_type = row_data.get("institution_type", "")
                color = self.get_color_for_value(inst_type, "institution_type")
                if color:
                    worksheet.cell(row=row_idx, column=col_idx).fill = self.colors[
                        color
                    ]
                col_idx += 1

                # Data year
                year = safe_int(row_data.get("data_year", 0))
                color = self.get_color_for_value(year, "data_year")
                if color:
                    worksheet.cell(row=row_idx, column=col_idx).fill = self.colors[
                        color
                    ]
                col_idx += 1

                # Percent female
                pct_female = safe_float(row_data.get("percent_female_undergrad", 0))
                color = self.get_color_for_value(pct_female, "percent_female")
                if color:
                    worksheet.cell(row=row_idx, column=col_idx).fill = self.colors[
                        color
                    ]
                worksheet.cell(row=row_idx, column=col_idx).number_format = "0.00%"
                col_idx += 2  # Skip total undergrads and grads

                # Undergrad/grad ratio
                ratio = safe_float(row_data.get("undergrad_grad_ratio", 0))
                color = self.get_color_for_value(ratio, "undergrad_grad_ratio")
                if color:
                    worksheet.cell(row=row_idx, column=col_idx).fill = self.colors[
                        color
                    ]
                col_idx += 4  # Skip other columns to acceptance rates

                # Acceptance rates (male, female, net, instate, outstate)
                for rate_key in [
                    "male_acceptance_rate",
                    "female_acceptance_rate",
                    "net_acceptance_rate",
                    "instate_acceptance_rate",
                    "outstate_acceptance_rate",
                ]:
                    rate = safe_float(row_data.get(rate_key, 0))
                    color = self.get_color_for_value(rate, "acceptance_rate")
                    if color:
                        worksheet.cell(row=row_idx, column=col_idx).fill = self.colors[
                            color
                        ]
                    worksheet.cell(row=row_idx, column=col_idx).number_format = "0.00%"
                    col_idx += 1

                # Yield rates
                for rate_key in [
                    "male_yield_rate",
                    "female_yield_rate",
                    "net_yield_rate",
                ]:
                    rate = safe_float(row_data.get(rate_key, 0))
                    color = self.get_color_for_value(rate, "yield_rate")
                    if color:
                        worksheet.cell(row=row_idx, column=col_idx).fill = self.colors[
                            color
                        ]
                    worksheet.cell(row=row_idx, column=col_idx).number_format = "0.00%"
                    col_idx += 1

                col_idx += 2  # Skip instate/outstate yield

                # SAT/ACT policy
                policy = row_data.get("sat_act_policy", "")
                color = self.get_color_for_value(policy, "sat_act_policy")
                if color:
                    worksheet.cell(row=row_idx, column=col_idx).fill = self.colors[
                        color
                    ]
                col_idx += 1

                # SAT scores
                sat_scores = [
                    "sat_composite_50th",
                    "sat_math_50th",
                    "sat_reading_50th",
                    "sat_composite_75th",
                    "sat_math_75th",
                    "sat_reading_75th",
                ]
                for i, score_key in enumerate(sat_scores):
                    score = safe_int(row_data.get(score_key, 0))
                    if score > 0:  # Only color if we have a valid score
                        if "composite" in score_key:
                            color = self.get_color_for_value(score, "sat_score")
                        else:
                            color = self.get_color_for_value(score, "sat_section")
                        if color:
                            worksheet.cell(row=row_idx, column=col_idx).fill = (
                                self.colors[color]
                            )
                    col_idx += 1

                # Early Action/Decision
                ea_available = row_data.get("early_action_available", "no")
                color = self.get_color_for_value(ea_available, "yes_no_green")
                if color:
                    worksheet.cell(row=row_idx, column=col_idx).fill = self.colors[
                        color
                    ]
                col_idx += 1

                ed_available = row_data.get("early_decision_available", "no")
                color = self.get_color_for_value(ed_available, "yes_no_green")
                if color:
                    worksheet.cell(row=row_idx, column=col_idx).fill = self.colors[
                        color
                    ]
                col_idx += 1

                # ED/EA acceptance rates
                if ed_available == "yes":
                    ed_rate = safe_float(row_data.get("ed_acceptance_rate", 0))
                    color = self.get_color_for_value(ed_rate, "acceptance_rate")
                    if color:
                        worksheet.cell(row=row_idx, column=col_idx).fill = self.colors[
                            color
                        ]
                    worksheet.cell(row=row_idx, column=col_idx).number_format = "0.00%"
                col_idx += 1

                if ea_available == "yes":
                    ea_rate = safe_float(row_data.get("ea_acceptance_rate", 0))
                    color = self.get_color_for_value(ea_rate, "acceptance_rate")
                    if color:
                        worksheet.cell(row=row_idx, column=col_idx).fill = self.colors[
                            color
                        ]
                    worksheet.cell(row=row_idx, column=col_idx).number_format = "0.00%"
                col_idx += 1

                # Double majoring
                double_major = row_data.get("double_majoring_allowed", "no")
                color = self.get_color_for_value(double_major, "yes_no_green")
                if color:
                    worksheet.cell(row=row_idx, column=col_idx).fill = self.colors[
                        color
                    ]
                col_idx += 1

                # Tuition
                tuition = safe_float(row_data.get("tuition_room_board", 0))
                color = self.get_color_for_value(tuition, "tuition")
                if color:
                    worksheet.cell(row=row_idx, column=col_idx).fill = self.colors[
                        color
                    ]

        print(f"Excel file created: {output_path}")


def main():
    # Configuration
    API_KEY = os.getenv("CLAUDE_API_KEY")
    if not API_KEY:
        print("Error: Please set the CLAUDE_API_KEY environment variable")
        print("You can set it temporarily in PowerShell using:")
        print('$env:CLAUDE_API_KEY="your-api-key-here"')
        return

    PDF_FOLDER = "CDS PDFs"  # Relative path to the PDF folder
    OUTPUT_FILE = "College Data Spreadsheet(s)/university_cds_analysis.xlsx"

    # Create output directory if it doesn't exist
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

    # Initialize analyzer
    analyzer = CDSAnalyzer(API_KEY, PDF_FOLDER)

    # Process PDFs
    print("Starting CDS analysis...")
    analyzer.process_pdfs()

    # Create Excel file
    print("Creating Excel file...")
    analyzer.create_excel_file(OUTPUT_FILE)

    print("Analysis complete!")


if __name__ == "__main__":
    main()

